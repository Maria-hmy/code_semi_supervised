import numpy as np
import sys
from exams.exam_genkyst import exam_genkyst
import matplotlib.pyplot as plt
from manage.manage_genkyst import genksyt_split
from openpyxl import load_workbook
from current import current_xlsx

output = '../../../../Desktop/'
T2_statistics = False
CT_statistics = True

if T2_statistics:
    modality = 'T2'
elif CT_statistics:
    modality = 'CT'
prefix = modality + '-LV-'
        
wb = load_workbook(current_xlsx()[0])
sh = wb['main']
list_ids = [sh.cell(row=rownum, column=1).value for rownum in range(2,sh.max_row+1)]
list_series = [sh.cell(row=rownum, column=8).value for rownum in range(2,sh.max_row+1)]

LV_volume, pkd, mayo = list(), list(), list()
ids_, patients, sexs, ages = list(), list(), list(), list()

for idx, id_ in enumerate(list_ids):
    print(id_)
    serie = int(list_series[idx])
    exam = exam_genkyst(int(id_), serie, modality, False)
    
    if CT_statistics:
        if exam.CT_exist and exam.liv_annot:
            exam = exam_genkyst(int(id_), serie, modality)
            if exam.id not in ids_:
                patients.append(exam.id)
                sexs.append(exam.sex)
                ages.append(exam.age)
                if exam.mutation == 'PKD1NT':
                    pkd.append(1)
                elif exam.mutation == 'PKD1T':
                    pkd.append(2)
                elif exam.mutation == 'PKD2':
                    pkd.append(3)
                elif exam.mutation == 'other':
                    pkd.append(4)
                else:
                    pkd.append(0)
            ids_.append(exam.id)
            if exam.mayo == '1A':
                mayo.append(0)
            elif exam.mayo == '1B':
                mayo.append(1)
            elif exam.mayo == '1C':
                mayo.append(2)
            elif exam.mayo == '1D':
                mayo.append(3)  
            elif exam.mayo == '1E':
                mayo.append(4) 
            elif exam.mayo == '2A':
                mayo.append(5)
            elif exam.mayo == '2B':
                mayo.append(6)
            xspacing = abs(exam.CT.get_qform()[0,0])
            yspacing = abs(exam.CT.get_qform()[1,1])
            zspacing = abs(exam.CT.get_qform()[2,2])
            spacing = xspacing*yspacing*zspacing
            LV_volume.append(float(spacing)*len(np.where(exam.LV.get_fdata()>0)[0])/1000.)
    elif T2_statistics:
        if exam.T2_exist and exam.liv_annot:
            exam = exam_genkyst(int(id_), serie, modality)
            if exam.id not in ids_:
                patients.append(exam.id)
                sexs.append(exam.sex)
                ages.append(exam.age)
                if exam.mutation == 'PKD1NT':
                    pkd.append(1)
                elif exam.mutation == 'PKD1T':
                    pkd.append(2)
                elif exam.mutation == 'PKD2':
                    pkd.append(3)                
                elif exam.mutation == 'other':
                    pkd.append(4)
                else:
                    pkd.append(0)
            ids_.append(exam.id)
            if exam.mayo == '1A':
                mayo.append(0)
            elif exam.mayo == '1B':
                mayo.append(1)
            elif exam.mayo == '1C':
                mayo.append(2)
            elif exam.mayo == '1D':
                mayo.append(3)  
            elif exam.mayo == '1E':
                mayo.append(4) 
            elif exam.mayo == '2A':
                mayo.append(5)
            elif exam.mayo == '2B':
                mayo.append(6)
            xspacing = abs(exam.T2.get_qform()[0,0])
            yspacing = abs(exam.T2.get_qform()[1,1])
            zspacing = abs(exam.T2.get_qform()[2,2])
            spacing = xspacing*yspacing*zspacing
            LV_volume.append(float(spacing)*len(np.where(exam.LV.get_fdata()>0)[0])/1000.)
    else:
        print('wrong setting!')

pkd, mayo = np.array(pkd), np.array(mayo)

with open(output+prefix+'statistics.txt', 'w') as file:
    file.write('number of examinations: %d\n'%len(ids_))
    file.write('number of patient: %d\n'%len(np.unique(np.array(patients))))
    file.write('number of man: %d\n'%len(np.where(np.array(sexs)=='M')[0]))
    file.write('number of woman: %d\n'%len(np.where(np.array(sexs)=='F')[0]))
    file.write('average age: %.2f\n'%np.mean(np.array(ages)))
    file.write('standard deviation age: %.2f\n'%(np.std(np.array(ages))))
    file.write('pourcentage of 1A: %.2f'%(len(np.where(mayo==0)[0])*100/mayo.shape[0])+' %\n')
    file.write('pourcentage of 1B: %.2f'%(len(np.where(mayo==1)[0])*100/mayo.shape[0])+' %\n')
    file.write('pourcentage of 1C: %.2f'%(len(np.where(mayo==2)[0])*100/mayo.shape[0])+' %\n')
    file.write('pourcentage of 1D: %.2f'%(len(np.where(mayo==3)[0])*100/mayo.shape[0])+' %\n')
    file.write('pourcentage of 1E: %.2f'%(len(np.where(mayo==4)[0])*100/mayo.shape[0])+' %\n')
    file.write('pourcentage of 2A: %.2f'%(len(np.where(mayo==5)[0])*100/mayo.shape[0])+' %\n')
    file.write('pourcentage of 2B: %.2f'%(len(np.where(mayo==6)[0])*100/mayo.shape[0])+' %\n')
    file.write('pourcentage of PKD1NT mutation: %.2f'%(len(np.where(pkd==1)[0])*100/pkd.shape[0])+' %\n')
    file.write('pourcentage of PKD1T mutation: %.2f'%(len(np.where(pkd==2)[0])*100/pkd.shape[0])+' %\n')
    file.write('pourcentage of PKD2 mutation: %.2f'%(len(np.where(pkd==3)[0])*100/pkd.shape[0])+' %\n')
    file.write('pourcentage of other mutation: %.2f'%(len(np.where(pkd==4)[0])*100/pkd.shape[0])+' %\n')
    file.write('pourcentage of unknown mutation: %.2f'%(len(np.where(pkd==0)[0])*100/pkd.shape[0])+' %\n')

