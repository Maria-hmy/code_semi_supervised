#!/usr/bin/env python3
# -*- coding: utf-8 -*-



import nibabel
import logging
from utils.utils import normalization_imgs, mask_zero
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from current import current_xlsx

class exam_genkyst:

    def __init__(self, id_, serie, modality, upload=True, load_info=True):
        self.id = id_
        self.serie = serie
        self.modality = modality

        if load_info:
            self.retrieve_info()

        if upload:
            self.load_images()  


            
    def retrieve_info(self):
        wb = load_workbook(self.xlsx)
        sh = wb['main']
        list_ids = [sh.cell(row=rownum, column=1).value for rownum in range(2, sh.max_row + 1)]
        rowindexes = np.where(np.array(list_ids)==self.id)[0]+2
        rowindex = rowindexes[0]
        if len(rowindexes)>1:
            list_series = [sh.cell(row=r, column=8).value for r in rowindexes]
            rowindex = rowindexes[np.where(np.array(list_series, dtype=int) == self.serie)[0]][0]
        row = [sh.cell(row=rowindex, column=col).value for col in range(1, sh.max_column + 1)]
        row = np.array(row)
        self.center    = str(row[1])
        self.sex       = str(row[3])
        self.size      = -1 if '?' in str(row[4]) else float(row[4])
        date1, date2 = row[2], row[5]
        self.age = -1
        if isinstance(date1, datetime) and isinstance(date2, datetime):
            self.age   = date2.year - date1.year - ((date2.month, date2.day) < (date1.month, date1.day))
        self.T2_exist  = True if 'x' in str(row[8]) else False
        self.CT_exist  = True if 'x' in str(row[11]) else False
        self.longi_MR  = 1 if row[12] == None else int(float(row[12]))
        self.longi_CT  = 1 if row[13] == None else int(float(row[13]))
        self.mayo      = str(row[14])
        self.mutation  = str(row[15])
        self.kid_annot = True if 'x' in str(row[16]) else False
        self.liv_annot = True if 'x' in str(row[17]) else False
        self.cancel    = True if 'x' in str(row[18]) else False

    def exam_upload(self, modality):
        if modality == 'T2' and self.T2_exist:
            self.T2 = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d-T2.nii.gz'%self.serie))
        elif modality == 'CT' and self.CT_exist:   
            self.CT = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d-CT.nii.gz'%self.serie))   
        if self.kid_annot:
            self.RK = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d-RK.nii.gz'%self.serie))
            self.LK = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d-LK.nii.gz'%self.serie))
            self.BK = mask_zero(self.RK)
            self.BK.get_fdata()[:,:,:] = self.RK.get_fdata()[:,:,:]+self.LK.get_fdata()[:,:,:]
        if self.liv_annot:
            self.LV = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d-LV.nii.gz'%self.serie))
        
    def normalize(self, modality):
        if modality == 'T2':
            self.T2.get_fdata()[:,:,:] = normalization_imgs(self.T2.get_fdata())[:,:,:]
        elif modality == 'CT':
            self.CT.get_fdata()[:,:,:] = normalization_imgs(self.CT.get_fdata())[:,:,:]
            
    def print_info(self):
        logging.basicConfig(level=logging.INFO, format='\n %(levelname)s: %(message)s')
        logging.info(f'''exam {self.id} uploaded:
        serie:         {self.serie}
        center:        {self.center}
        sex:           {self.sex}
        age:           {self.age}
        size:          {self.size}      
        T2:            {self.T2_exist}
        CT:            {self.CT_exist}
        mayo:          {self.mayo}
        mutation:      {self.mutation}
        ''')