#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import nibabel
import logging
from utils.utils import normalization_imgs
import numpy as np
from openpyxl import load_workbook

class exam_berlin:

    def __init__(self, id_, serie, modality, upload=True):
        self.root = './../../data/berlin/'
        self.xlsx = self.root+'2024-05-05-berlin-dataset.xlsx'
        self.folder = self.root+'dataset/'
        self.id = '%0*d'%(3,id_)
        self.serie = serie
        self.retrieve_info()
        if upload:
            self.exam_upload(modality)
            self.print_info()
            
    def retrieve_info(self):
        wb = load_workbook(self.xlsx, data_only=True)
        sh = wb['main'] 
        list_ids = [sh.cell(row=rownum, column=1).value for rownum in range(2, sh.max_row + 1)]
        rowindexes = np.where(np.array(list_ids) == self.id)[0] + 2
        rowindex = rowindexes[0]
        if len(rowindexes) > 1:
            list_series = [sh.cell(row=r, column=6).value for r in rowindexes]  # Assuming series is in column F
            rowindex = rowindexes[np.where(np.array(list_series, dtype=int) == self.serie)[0]][0]
        row = [sh.cell(row=rowindex, column=col).value for col in range(1, sh.max_column + 1)]
        row = np.array(row)
        self.T2c_exist = True if 'x' in row[6] else False
        self.T2a_exist = True if 'x' in row[7] else False
        self.CT_exist  = True if 'x' in row[8] else False

    def exam_upload(self, modality):
        if modality == 'T2c' and self.T2c_exist:
            self.T2c = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d.nii.gz'%self.serie))
        elif modality == 'T2a' and self.T2a_exist:   
            self.T2a = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d.nii.gz'%self.serie))  
        elif modality == 'CT' and self.CT_exist:   
            self.CT = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d.nii.gz'%self.serie))   
        self.LV = nibabel.as_closest_canonical(nibabel.load(self.folder+self.id+'-%02d-LV.nii.gz'%self.serie))
        
    def normalize(self, modality):
        if modality == 'T2c':
            self.T2c.get_fdata()[:,:,:] = normalization_imgs(self.T2c.get_fdata())[:,:,:]
        if modality == 'T2a':
            self.T2a.get_fdata()[:,:,:] = normalization_imgs(self.T2a.get_fdata())[:,:,:]
        elif modality == 'CT':
            self.CT.get_fdata()[:,:,:] = normalization_imgs(self.CT.get_fdata())[:,:,:]
            
    def print_info(self):
        logging.basicConfig(level=logging.INFO, format='\n %(levelname)s: %(message)s')
        logging.info(f'''exam {self.id} uploaded:
        serie:         {self.serie}
        T2c:           {self.T2c_exist}
        T2a:           {self.T2a_exist}
        CT:            {self.CT_exist}
        ''')